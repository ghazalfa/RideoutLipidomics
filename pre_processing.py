import csv
from openpyxl import load_workbook 
from openpyxl import Workbook
from openpyxl import load_workbook, Workbook

# Load the workbook and create a new one
wb = load_workbook(filename='/Users/ghazalfallahpour/Desktop/20221004_FlyBrain_Lipids_Summary.xlsx')
wb1 = Workbook()

# Select the third sheet in the original workbook
sheet = wb.worksheets[2]

# Define the start and end rows and columns in the dataset
start_row = 2
end_row = 1350
start_col = 4
end_col = 29

# Dictionary to store total values
total = {}

# Iterate through rows and cells in the specified range
for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
    for cell in row[start_col-1:end_col]:
        
        # Check if the cell is in the "metaName" column (column 4)
        if cell.col_idx == 4:
            if 'w/o' in cell.value or 'Unknown' in cell.value or 'RIKEN' in cell.value:
                state = False
            else:
                state = True
                
            if state:
                metaName = cell.value
                if total.get(metaName) is None:
                    # Update metabolite name to each dictionary
                    total.update({metaName: [0] * 24})
        
        # If the state is true, update the values in the dictionary
        if state:
            if cell.col_idx > 5:
                allValues = total[metaName]
                specificValue = allValues[cell.col_idx - 6]
                specificValue += cell.value
                allValues[cell.col_idx - 6] = specificValue
                total.update({metaName: allValues})

# Load the workbook again
wb = load_workbook(filename='/Users/ghazalfallahpour/Desktop/20221004_FlyBrain_Lipids_Summary.xlsx')

# Select the active sheet
sheet = wb.active

# Define the start and end rows and columns
start_row = 2
end_row = 4200
start_col = 4
end_col = 29

# Dictionary to store totalNeg values
totalNeg = {}

# Iterate through rows and cells in the specified range
for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
    for cell in row[start_col-1:end_col]:
        
        # Check if the cell is in the "metaName" column (column 4)
        if cell.col_idx == 4:
            if 'w/o' in cell.value or 'Unknown' in cell.value or 'RIKEN' in cell.value:
                state = False
            else:
                state = True
                
            if state:
                metaName = cell.value
                if totalNeg.get(metaName) is None:
                    # Update metabolite name to each dictionary
                    totalNeg.update({metaName: [0] * 24})
        
        # If the state is true, update the values in the dictionary
        if state:
            if cell.col_idx > 5:
                allValues = totalNeg[metaName]
                specificValue = allValues[cell.col_idx - 6]
                specificValue += cell.value
                allValues[cell.col_idx - 6] = specificValue
                totalNeg.update({metaName: allValues})

# Update total dictionary with averaged values from total and totalNeg
for metaName in totalNeg:
    counter = 0
    while counter < 24:
        sumValues = total[metaName][counter] + totalNeg[metaName][counter]
        avgValue = sumValues / 2
        total[metaName][counter] = avgValue
        counter += 1

        
        
        
        
        
        
